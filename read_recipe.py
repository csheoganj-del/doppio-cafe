import openpyxl

path = r"C:\Users\MASTER PC\Downloads\new item recipe copy (1).xlsx"

try:
    wb = openpyxl.load_workbook(path)
    for sheet in wb.sheetnames:
        print(f"--- SHEET: {sheet} ---")
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            # Print non-empty rows nicely
            if any(cell is not None for cell in row):
                print(row)
except Exception as e:
    print("Error reading with openpyxl:", e)
