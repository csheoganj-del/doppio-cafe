import openpyxl
import json

path = r"C:\Users\MASTER PC\Downloads\new item recipe copy (1).xlsx"
wb = openpyxl.load_workbook(path)

recipes_data = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}")
    
    current_category = None
    
    for row in ws.iter_rows(values_only=True):
        if not row or not any(cell is not None for cell in row):
            continue
            
        first_cell = row[0]
        
        # If the row has only one cell filled, it's a sub-category or header
        filled_cells = [c for c in row if c is not None]
        if len(filled_cells) == 1 and isinstance(first_cell, str) and not first_cell.endswith(' '):
            current_category = first_cell.strip()
            continue
            
        if first_cell and isinstance(first_cell, str):
            item_name = first_cell.strip()
            
            # Map ingredients
            ingredients = {}
            # Go through pairs of (ingredient, amount)
            for i in range(1, len(row), 2):
                if i+1 < len(row):
                    ing = row[i]
                    amt = row[i+1]
                    if ing and amt:
                        ing_clean = str(ing).strip().lower().replace(' ', '_')
                        ingredients[ing_clean] = str(amt).strip()
            
            recipes_data[item_name.lower()] = {
                "original_name": item_name,
                "category": current_category or sheet_name,
                "ingredients": ingredients
            }

# Dump to JSON
with open('recipes.json', 'w', encoding='utf-8') as f:
    json.dump(recipes_data, f, indent=2, ensure_ascii=False)

print(f"Successfully exported {len(recipes_data)} recipes to recipes.json!")
